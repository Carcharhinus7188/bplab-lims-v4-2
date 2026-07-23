from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import COMPANY_CN, COMPANY_EN, OUTPUT_DIR
from .db import query
from .services import build_report_snapshot


HEADER_FILL = PatternFill("solid", fgColor="0B63CE")
SUB_FILL = PatternFill("solid", fgColor="DDEBFA")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = max(12, min(40, max(len(str(ws.cell(row, col).value or "")) for row in range(1, ws.max_row + 1)) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_trace_excel(
    commission_id: int,
    output_dir: Path | None = None,
    db_path: Path | None = None,
) -> Path:
    snapshot = build_report_snapshot(commission_id, db_path)
    commission = snapshot["commission"]
    attachments = query(
        """SELECT a.*,u.display_name AS uploader,t.task_no
           FROM attachments a
           LEFT JOIN users u ON u.id=a.uploaded_by
           LEFT JOIN tasks t ON t.id=a.task_id
           WHERE a.commission_id=? ORDER BY a.id""",
        (commission_id,),
        db_path,
    )
    audits = query(
        """SELECT l.*,u.display_name AS actor
           FROM audit_log l LEFT JOIN users u ON u.id=l.actor_id
           WHERE (l.entity_type='commission' AND l.entity_id=?)
              OR (l.entity_type='task' AND l.entity_id IN (
                    SELECT CAST(id AS TEXT) FROM tasks WHERE commission_id=?
                 ))
              OR (l.entity_type='report' AND l.entity_id IN (
                    SELECT CAST(id AS TEXT) FROM reports WHERE commission_id=?
                 ))
           ORDER BY l.id""",
        (str(commission_id), commission_id, commission_id),
        db_path,
    )
    wb = Workbook()
    guide = wb.active
    guide.title = "使用说明"
    guide.append(["项目", "内容"])
    guide.append(["公司", f"{COMPANY_CN} / {COMPANY_EN}"])
    guide.append(["委托编号", commission["commission_no"]])
    guide.append(["文件定位", "内部实验数据与附件追溯工作簿；不替代正式原始记录和检验报告。"])
    guide.append(["附件原则", "图片、截图、曲线、原始数据文件只在本工作簿建立索引，不塞入受控原始记录。"])
    guide.append(["完整性", "附件以SHA-256校验值识别；文件内容变化会导致校验值变化。"])

    ledger = wb.create_sheet("实验总台账")
    ledger.append(["任务编号", "样品编号", "实验名称", "方法", "地点", "实验员", "开始时间", "提交时间", "复核时间", "判定", "配置版本/快照"])
    for task in snapshot["tasks"]:
        ledger.append(
            [
                task["task_no"],
                "、".join(task["inherited"].get("sample_ids", [])),
                task["config"].get("name", ""),
                task["config"].get("method", ""),
                task.get("location", ""),
                commission.get("tester_name", ""),
                task.get("started_at", ""),
                task.get("submitted_at", ""),
                task.get("reviewed_at", ""),
                task["calculations"].get("judgment", ""),
                task["config"].get("version", "1.0"),
            ]
        )

    index = wb.create_sheet("附件索引")
    index.append(["附件编号", "任务编号", "样品编号", "附件类型", "原文件名", "相对路径", "生成/拍摄时间", "上传人", "内容说明", "原始/处理关系", "SHA-256", "复核状态"])
    for item in attachments:
        index.append(
            [
                item["attachment_no"],
                item.get("task_no", ""),
                item.get("sample_id", ""),
                item["attachment_type"],
                item["original_filename"],
                item["relative_path"],
                item["generated_at"],
                item.get("uploader", ""),
                item.get("description", ""),
                item.get("source_relation", ""),
                item["sha256"],
                item["review_status"],
            ]
        )

    images = wb.create_sheet("图片粘贴区")
    images.append(["附件编号", "任务/样品", "图片说明", "粘贴位置"])
    for item in attachments:
        if item["attachment_type"] in {"实验过程照片", "图像文件", "软件截图", "异常证明文件"}:
            images.append([item["attachment_no"], f"{item.get('task_no','')} / {item.get('sample_id','')}", item.get("description", ""), "在此单元格附近粘贴经核对的图片"])

    changes = wb.create_sheet("修改记录")
    changes.append(["时间", "操作人", "操作", "对象", "对象编号", "详细信息"])
    for item in audits:
        changes.append([item["created_at"], item.get("actor", "系统"), item["action"], item["entity_type"], item["entity_id"], item["detail_json"]])

    reviews = wb.create_sheet("复核记录")
    reviews.append(["任务编号", "实验名称", "数据快照SHA-256", "提交状态", "复核时间", "复核员", "结论"])
    for task in snapshot["tasks"]:
        versions = query(
            "SELECT * FROM record_versions WHERE task_id=? ORDER BY version",
            (task["id"],),
            db_path,
        )
        for version in versions:
            reviews.append(
                [
                    task["task_no"],
                    task["config"].get("name", ""),
                    version["data_hash"],
                    version["status"],
                    task.get("reviewed_at", ""),
                    commission.get("reviewer_name", ""),
                    task["calculations"].get("judgment", ""),
                ]
            )

    for task in snapshot["tasks"]:
        title = task["config"].get("name", task["experiment_code"])[:28]
        if title in wb.sheetnames:
            title = f"{title}-{task['id']}"
        ws = wb.create_sheet(title)
        ws.append(["字段", "内容"])
        ws.append(["任务编号", task["task_no"]])
        ws.append(["样品编号", "、".join(task["inherited"].get("sample_ids", []))])
        ws.append(["检测方法", task["config"].get("method", "")])
        ws.append(["标准要求", task["calculations"].get("standard_requirement", "")])
        ws.append(["实际结果", task["calculations"].get("report_result", "")])
        ws.append(["单项结论", task["calculations"].get("judgment", "")])
        ws.append(["环境数据", json.dumps(task["data"].get("environment", {}), ensure_ascii=False)])
        ws.append(["参数", json.dumps(task["data"].get("parameters", {}), ensure_ascii=False)])
        ws.append(["过程数据", json.dumps(task["data"].get("run", {}), ensure_ascii=False)])
        ws.append(["逐样原始数据与计算", json.dumps(task["calculations"].get("samples", []), ensure_ascii=False)])

    for ws in wb.worksheets:
        _style_sheet(ws)
    target_dir = output_dir or OUTPUT_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{commission['commission_no']}_内部实验数据追溯工作簿.xlsx"
    wb.save(target)
    return target
